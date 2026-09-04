import csv
import io
import re
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from app.database import get_db
from app.models.user import User
from app.models.workspace import Workspace
from app.models.student_directory import StudentDirectory, DirectoryStudent
from app.schemas.student_directory import (
    StudentDirectoryCreate, StudentDirectoryUpdate, StudentDirectoryResponse,
    DirectoryStudentCreate, DirectoryStudentUpdate, DirectoryStudentResponse,
    CSVImportResult
)
from app.utils.security import get_current_user
from app.services.workspace_service import get_current_workspace
from app.utils.tabular_parser import (
    parse_tabular_file,
    generate_student_template_csv,
    generate_student_template_excel
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/student-directories", tags=["student_directories"])

# ──────── REFERENCE TEMPLATES ────────

@router.get("/template/csv")
def download_roster_csv_template():
    """Returns a downloadable sample CSV template for bulk student roster import."""
    csv_content = generate_student_template_csv()
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=student_roster_template.csv"}
    )

@router.get("/template/excel")
@router.get("/template/xlsx")
def download_roster_excel_template():
    """Returns a downloadable styled Excel (.xlsx) reference template for bulk student roster import."""
    excel_bytes = generate_student_template_excel()
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_roster_template.xlsx"}
    )

def _get_authorized_directory(directory_id: str, workspace: Workspace, db: Session) -> StudentDirectory:
    directory = db.query(StudentDirectory).filter(
        StudentDirectory.id == directory_id,
        StudentDirectory.workspace_id == workspace.id,
        StudentDirectory.is_deleted == False
    ).first()
    if not directory:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student directory not found or does not belong to the active workspace"
        )
    return directory

@router.get("", response_model=List[StudentDirectoryResponse])
@router.get("/", response_model=List[StudentDirectoryResponse])
def list_student_directories(
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Lists all student directories belonging to the active workspace."""
    directories = db.query(StudentDirectory).filter(
        StudentDirectory.workspace_id == current_workspace.id,
        StudentDirectory.is_deleted == False
    ).order_by(StudentDirectory.created_at.desc()).all()

    res = []
    for d in directories:
        count = db.query(DirectoryStudent).filter(
            DirectoryStudent.directory_id == d.id,
            DirectoryStudent.is_deleted == False
        ).count()
        res.append({
            "id": d.id,
            "workspace_id": d.workspace_id,
            "name": d.name,
            "description": d.description,
            "created_by": d.created_by,
            "student_count": count,
            "is_active": d.is_active,
            "created_at": d.created_at
        })
    return res

@router.post("", response_model=StudentDirectoryResponse)
@router.post("/", response_model=StudentDirectoryResponse)
def create_student_directory(
    payload: StudentDirectoryCreate,
    current_workspace: Workspace = Depends(get_current_workspace),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Creates a new student directory in the active workspace."""
    directory = StudentDirectory(
        workspace_id=current_workspace.id,
        name=payload.name.strip(),
        description=payload.description.strip() if payload.description else None,
        created_by=current_user.id,
        is_active=True
    )
    db.add(directory)
    db.flush()

    student_count = 0
    if payload.initial_students:
        for s in payload.initial_students:
            if not s.name or not s.name.strip():
                continue
            email_val = s.email.strip().lower() if s.email and s.email.strip() else None
            roll_val = s.roll_number.strip() if s.roll_number and s.roll_number.strip() else None
            phone_val = s.phone.strip() if s.phone and s.phone.strip() else None
            student = DirectoryStudent(
                directory_id=directory.id,
                name=s.name.strip(),
                email=email_val,
                roll_number=roll_val,
                phone=phone_val,
                student_code=s.student_code,
                status=s.status or "active"
            )
            db.add(student)
            student_count += 1

    db.commit()
    db.refresh(directory)

    return {
        "id": directory.id,
        "workspace_id": directory.workspace_id,
        "name": directory.name,
        "description": directory.description,
        "created_by": directory.created_by,
        "student_count": student_count,
        "is_active": directory.is_active,
        "created_at": directory.created_at
    }

@router.get("/{directory_id}", response_model=StudentDirectoryResponse)
def get_student_directory(
    directory_id: str,
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Gets details of a student directory."""
    directory = _get_authorized_directory(directory_id, current_workspace, db)
    count = db.query(DirectoryStudent).filter(
        DirectoryStudent.directory_id == directory.id,
        DirectoryStudent.is_deleted == False
    ).count()
    return {
        "id": directory.id,
        "workspace_id": directory.workspace_id,
        "name": directory.name,
        "description": directory.description,
        "created_by": directory.created_by,
        "student_count": count,
        "is_active": directory.is_active,
        "created_at": directory.created_at
    }

@router.put("/{directory_id}", response_model=StudentDirectoryResponse)
def update_student_directory(
    directory_id: str,
    payload: StudentDirectoryUpdate,
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Updates directory metadata."""
    directory = _get_authorized_directory(directory_id, current_workspace, db)
    if payload.name is not None:
        directory.name = payload.name.strip()
    if payload.description is not None:
        directory.description = payload.description.strip()
    db.commit()
    db.refresh(directory)
    count = db.query(DirectoryStudent).filter(
        DirectoryStudent.directory_id == directory.id,
        DirectoryStudent.is_deleted == False
    ).count()
    return {
        "id": directory.id,
        "workspace_id": directory.workspace_id,
        "name": directory.name,
        "description": directory.description,
        "created_by": directory.created_by,
        "student_count": count,
        "is_active": directory.is_active,
        "created_at": directory.created_at
    }

@router.delete("/{directory_id}")
def delete_student_directory(
    directory_id: str,
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Soft deletes a student directory."""
    directory = _get_authorized_directory(directory_id, current_workspace, db)
    directory.is_deleted = True
    db.commit()
    return {"message": "Directory deleted successfully"}

# ──────── STUDENTS INSIDE DIRECTORY ────────

@router.get("/{directory_id}/students", response_model=List[DirectoryStudentResponse])
def list_directory_students(
    directory_id: str,
    search: Optional[str] = None,
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Lists students in the directory with optional search."""
    directory = _get_authorized_directory(directory_id, current_workspace, db)
    query = db.query(DirectoryStudent).filter(
        DirectoryStudent.directory_id == directory.id,
        DirectoryStudent.is_deleted == False
    )
    if search:
        s = f"%{search.strip()}%"
        query = query.filter(
            (DirectoryStudent.name.ilike(s)) |
            (DirectoryStudent.email.ilike(s)) |
            (DirectoryStudent.roll_number.ilike(s))
        )
    return query.order_by(DirectoryStudent.name.asc()).all()

@router.post("/{directory_id}/students", response_model=DirectoryStudentResponse)
def add_directory_student(
    directory_id: str,
    payload: DirectoryStudentCreate,
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Adds a single student to the directory."""
    directory = _get_authorized_directory(directory_id, current_workspace, db)
    if not payload.name or not payload.name.strip():
        raise HTTPException(status_code=400, detail="Student name is required")

    email_val = payload.email.strip().lower() if payload.email and payload.email.strip() else None
    roll_val = payload.roll_number.strip() if payload.roll_number and payload.roll_number.strip() else None

    # Check for directory-level duplicate email or roll
    if email_val:
        dup = db.query(DirectoryStudent).filter(
            DirectoryStudent.directory_id == directory.id,
            DirectoryStudent.email == email_val,
            DirectoryStudent.is_deleted == False
        ).first()
        if dup:
            raise HTTPException(status_code=400, detail=f"A student with email '{email_val}' already exists in this directory")

    if roll_val:
        dup = db.query(DirectoryStudent).filter(
            DirectoryStudent.directory_id == directory.id,
            DirectoryStudent.roll_number == roll_val,
            DirectoryStudent.is_deleted == False
        ).first()
        if dup:
            raise HTTPException(status_code=400, detail=f"A student with roll number '{roll_val}' already exists in this directory")

    student = DirectoryStudent(
        directory_id=directory.id,
        name=payload.name.strip(),
        email=email_val,
        roll_number=roll_val,
        phone=payload.phone.strip() if payload.phone else None,
        student_code=payload.student_code,
        status=payload.status or "active"
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return student

@router.put("/{directory_id}/students/{student_id}", response_model=DirectoryStudentResponse)
def update_directory_student(
    directory_id: str,
    student_id: str,
    payload: DirectoryStudentUpdate,
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Updates a student's profile inside the directory."""
    directory = _get_authorized_directory(directory_id, current_workspace, db)
    student = db.query(DirectoryStudent).filter(
        DirectoryStudent.id == student_id,
        DirectoryStudent.directory_id == directory.id,
        DirectoryStudent.is_deleted == False
    ).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found in this directory")

    if payload.name is not None:
        student.name = payload.name.strip()
    if payload.email is not None:
        email_val = payload.email.strip().lower() if payload.email.strip() else None
        if email_val and email_val != student.email:
            dup = db.query(DirectoryStudent).filter(
                DirectoryStudent.directory_id == directory.id,
                DirectoryStudent.email == email_val,
                DirectoryStudent.id != student.id,
                DirectoryStudent.is_deleted == False
            ).first()
            if dup:
                raise HTTPException(status_code=400, detail=f"Another student with email '{email_val}' already exists")
        student.email = email_val
    if payload.roll_number is not None:
        roll_val = payload.roll_number.strip() if payload.roll_number.strip() else None
        if roll_val and roll_val != student.roll_number:
            dup = db.query(DirectoryStudent).filter(
                DirectoryStudent.directory_id == directory.id,
                DirectoryStudent.roll_number == roll_val,
                DirectoryStudent.id != student.id,
                DirectoryStudent.is_deleted == False
            ).first()
            if dup:
                raise HTTPException(status_code=400, detail=f"Another student with roll number '{roll_val}' already exists")
        student.roll_number = roll_val
    if payload.phone is not None:
        student.phone = payload.phone.strip() if payload.phone.strip() else None
    if payload.status is not None:
        student.status = payload.status

    db.commit()
    db.refresh(student)
    return student

@router.delete("/{directory_id}/students/{student_id}")
def delete_directory_student(
    directory_id: str,
    student_id: str,
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Deletes a student from the directory."""
    directory = _get_authorized_directory(directory_id, current_workspace, db)
    student = db.query(DirectoryStudent).filter(
        DirectoryStudent.id == student_id,
        DirectoryStudent.directory_id == directory.id,
        DirectoryStudent.is_deleted == False
    ).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found in this directory")
    student.is_deleted = True
    db.commit()
    return {"message": "Student removed from directory"}

# ──────── CSV & EXCEL IMPORT / EXPORT ────────

@router.post("/{directory_id}/import", response_model=CSVImportResult)
@router.post("/{directory_id}/import-csv", response_model=CSVImportResult)
def import_students_csv(
    directory_id: str,
    file: UploadFile = File(...),
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """
    Imports students from any CSV, TSV, or Excel file (.xlsx, .xls, .csv, .tsv, .txt) into the specified directory.
    Performs header normalization, whitespace trimming, email validation,
    duplicate checks, and detailed row error reporting.
    """
    directory = _get_authorized_directory(directory_id, current_workspace, db)

    valid_extensions = (".csv", ".xlsx", ".xls", ".tsv", ".txt", ".xlsm", ".xltx", ".xltm")
    if not file.filename.lower().endswith(valid_extensions):
        raise HTTPException(
            status_code=400,
            detail="Invalid file format. Please upload an Excel workbook (.xlsx, .xls) or CSV/TSV (.csv, .tsv) file."
        )

    file_bytes = file.file.read()
    try:
        rows = parse_tabular_file(file_bytes, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="The file is empty or contains no valid student records.")

    # Check that at least 'name' or 'email' is detected
    sample_keys = set().union(*(r.keys() for r in rows[:10]))
    if "name" not in sample_keys and "email" not in sample_keys:
        raise HTTPException(
            status_code=400,
            detail="File must contain at least a 'Name' or 'Email' column (e.g. Full Name, Email, Roll Number, Phone)"
        )

    # Fetch existing emails and roll numbers for this directory
    existing_emails = set(
        e[0].lower() for e in db.query(DirectoryStudent.email).filter(
            DirectoryStudent.directory_id == directory.id,
            DirectoryStudent.email.isnot(None),
            DirectoryStudent.is_deleted == False
        ).all()
    )
    existing_rolls = set(
        r[0] for r in db.query(DirectoryStudent.roll_number).filter(
            DirectoryStudent.directory_id == directory.id,
            DirectoryStudent.roll_number.isnot(None),
            DirectoryStudent.is_deleted == False
        ).all()
    )

    imported_count = 0
    skipped_count = 0
    errors = []
    seen_file_emails = set()
    seen_file_rolls = set()

    for row_idx, item in enumerate(rows, start=2):
        raw_name = item.get("name", "") or item.get("full_name", "")
        raw_email = item.get("email", "").lower() if item.get("email") else ""
        raw_roll = item.get("roll_number", "")
        raw_phone = item.get("phone", "")

        if not raw_name and not raw_email:
            skipped_count += 1
            errors.append(f"Row {row_idx}: Missing candidate name and email")
            continue

        name = raw_name or raw_email.split("@")[0].capitalize()
        email = raw_email if raw_email else None
        roll = raw_roll if raw_roll else None
        phone = raw_phone if raw_phone else None

        # Email format check
        if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            skipped_count += 1
            errors.append(f"Row {row_idx}: Invalid email format '{email}'")
            continue

        # Duplicate checks
        if email and (email in existing_emails or email in seen_file_emails):
            skipped_count += 1
            errors.append(f"Row {row_idx}: Duplicate email '{email}'")
            continue

        if roll and (roll in existing_rolls or roll in seen_file_rolls):
            skipped_count += 1
            errors.append(f"Row {row_idx}: Duplicate roll number '{roll}'")
            continue

        if email:
            seen_file_emails.add(email)
        if roll:
            seen_file_rolls.add(roll)

        student = DirectoryStudent(
            directory_id=directory.id,
            name=name,
            email=email,
            roll_number=roll,
            phone=phone,
            status="active"
        )
        db.add(student)
        imported_count += 1

    db.commit()

    return {
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "errors": errors
    }

@router.get("/{directory_id}/export")
@router.get("/{directory_id}/export-csv")
def export_students_csv(
    directory_id: str,
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Exports all students in the directory as a CSV file."""
    directory = _get_authorized_directory(directory_id, current_workspace, db)
    students = db.query(DirectoryStudent).filter(
        DirectoryStudent.directory_id == directory.id,
        DirectoryStudent.is_deleted == False
    ).order_by(DirectoryStudent.roll_number.asc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Roll Number", "Name", "Email", "Phone", "Status", "Joined Date"])

    for s in students:
        writer.writerow([
            s.roll_number or "",
            s.name,
            s.email or "",
            s.phone or "",
            s.status,
            s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else ""
        ])

    output.seek(0)
    filename = f"{re.sub(r'[^a-zA-Z0-9]+', '_', directory.name.lower())}_roster.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/{directory_id}/export-excel")
def export_students_excel(
    directory_id: str,
    current_workspace: Workspace = Depends(get_current_workspace),
    db: Session = Depends(get_db)
):
    """Exports all students in the directory as a formatted Excel (.xlsx) workbook."""
    directory = _get_authorized_directory(directory_id, current_workspace, db)
    students = db.query(DirectoryStudent).filter(
        DirectoryStudent.directory_id == directory.id,
        DirectoryStudent.is_deleted == False
    ).order_by(DirectoryStudent.roll_number.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{directory.name[:25]} Roster"
    header_fill = PatternFill(start_color="C84B18", end_color="C84B18", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="242321")

    headers = ["Roll Number", "Full Name", "Email Address", "Phone Number", "Status", "Joined Date"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for s in students:
        ws.append([
            s.roll_number or "",
            s.name,
            s.email or "",
            s.phone or "",
            s.status,
            s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else ""
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"{re.sub(r'[^a-zA-Z0-9]+', '_', directory.name.lower())}_roster.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
