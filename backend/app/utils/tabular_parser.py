"""
Universal Tabular Parser & Template Generator for EduQuizX.
Supports all CSV, TSV, TXT, and Excel (.xlsx, .xls, .xlsm) formats.
Handles delimiter sniffing, encoding fallbacks, header normalization,
whitespace cleaning, and reference sample file creation.
"""

import io
import os
import re
import csv
import logging
from typing import List, Dict, Any, Tuple, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Canonical field mappings for student/candidate rosters
HEADER_ALIASES = {
    "name": [
        "name", "full_name", "fullname", "student_name", "student", 
        "candidate_name", "candidate", "learner_name", "learner"
    ],
    "first_name": ["first_name", "firstname", "fname", "given_name"],
    "last_name": ["last_name", "lastname", "lname", "surname", "family_name"],
    "email": [
        "email", "student_email", "mail", "e_mail", "email_address", 
        "email_id", "user_email", "login_email"
    ],
    "roll_number": [
        "roll_number", "roll_no", "rollno", "roll", "candidate_id", 
        "student_id", "id", "reg_no", "registration_no", "enrollment_no", 
        "admission_no", "urn", "prn", "index_number", "roll_student_id", "student_roll_no"
    ],
    "phone": [
        "phone", "phone_number", "mobile", "mobile_no", "mobile_number", 
        "contact", "contact_no", "contact_number", "cell", "telephone", "whatsapp"
    ],
    "division": [
        "division", "div", "section", "sec", "batch", "class", "grade", "group", "division_section"
    ],
    "department": [
        "department", "dept", "department_name", "branch", "course", "program", "stream", "department_program"
    ]
}


def _clean_header_string(header: Any) -> str:
    """Sanitizes raw header into lowercase snake_case identifier."""
    if header is None:
        return ""
    text = str(header).strip().lower()
    return re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_")


def _map_columns(raw_columns: List[Any]) -> Tuple[Dict[str, str], bool]:
    """
    Maps raw column names to canonical fields ('name', 'email', 'roll_number', etc.).
    Returns (mapping, has_split_name) where mapping is {raw_col: canonical_field}.
    """
    mapping = {}
    has_first = False
    has_last = False

    cleaned_cols = [(_clean_header_string(c), c) for c in raw_columns]

    for clean_name, orig_col in cleaned_cols:
        matched = False
        for canon, aliases in HEADER_ALIASES.items():
            if clean_name in aliases:
                mapping[orig_col] = canon
                if canon == "first_name":
                    has_first = True
                elif canon == "last_name":
                    has_last = True
                matched = True
                break
        if not matched:
            mapping[orig_col] = clean_name

    has_split_name = has_first and has_last and "name" not in mapping.values()
    return mapping, has_split_name


def parse_tabular_file(file_bytes: bytes, filename: str) -> List[Dict[str, str]]:
    """
    Parses an uploaded file in any CSV or Excel format into a normalized list of dicts.
    Supported extensions: .csv, .xlsx, .xls, .xlsm, .tsv, .txt.
    """
    if not file_bytes:
        return []

    ext = os.path.splitext(filename.lower())[1]
    df: Optional[pd.DataFrame] = None

    # 1. EXCEL WORKBOOK PARSING (.xlsx, .xls, .xlsm)
    if ext in [".xlsx", ".xls", ".xlsm", ".xltx", ".xltm"]:
        try:
            # First try openpyxl / default engine
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
        except Exception as e:
            logger.warning(f"Default Excel engine failed for {filename}: {e}. Retrying with fallback engine...")
            try:
                # If .xls, try xlrd
                if ext == ".xls":
                    df = pd.read_excel(io.BytesIO(file_bytes), engine="xlrd", dtype=str)
                else:
                    df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl", dtype=str)
            except Exception as e2:
                raise ValueError(f"Failed to read Excel workbook '{filename}': {str(e2)}")

    # 2. DELIMITED TEXT / CSV PARSING (.csv, .tsv, .txt, or other)
    else:
        encodings_to_try = ["utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-1"]
        parsed = False
        last_error = None

        for enc in encodings_to_try:
            try:
                # pandas with sep=None and engine='python' sniffs comma, semicolon, tab, pipe automatically
                df = pd.read_csv(
                    io.BytesIO(file_bytes),
                    sep=None,
                    engine="python",
                    encoding=enc,
                    dtype=str,
                    skipinitialspace=True,
                    on_bad_lines="skip"
                )
                parsed = True
                break
            except Exception as err:
                last_error = err
                continue

        if not parsed:
            # Fallback to standard csv.reader with manual decoding
            try:
                text_content = file_bytes.decode("utf-8-sig", errors="replace")
                # Detect delimiter
                sample = text_content[:4096]
                delimiter = ","
                if "\t" in sample and sample.count("\t") > sample.count(","):
                    delimiter = "\t"
                elif ";" in sample and sample.count(";") > sample.count(","):
                    delimiter = ";"
                elif "|" in sample and sample.count("|") > sample.count(","):
                    delimiter = "|"

                lines = [l for l in text_content.splitlines() if l.strip()]
                if not lines:
                    return []
                reader = csv.reader(lines, delimiter=delimiter)
                header = next(reader, None)
                if not header:
                    return []
                records = []
                for row in reader:
                    rec = {}
                    for i, h in enumerate(header):
                        rec[h] = row[i] if i < len(row) else ""
                    records.append(rec)
                df = pd.DataFrame(records, dtype=str)
            except Exception as e3:
                raise ValueError(f"Could not parse delimited file '{filename}': {last_error or e3}")

    if df is None or df.empty:
        return []

    # Clean DataFrame: replace NaN with empty string
    df = df.fillna("")

    # Map headers to canonical names
    raw_columns = list(df.columns)
    col_mapping, has_split_name = _map_columns(raw_columns)

    normalized_rows = []
    for _, row in df.iterrows():
        # Check if entire row is whitespace
        values = [str(row[c]).strip() for c in raw_columns]
        if not any(values):
            continue

        item: Dict[str, str] = {}
        for orig_col, canon_col in col_mapping.items():
            val = str(row[orig_col]).strip()
            # Clean floating point artifact e.g. "123.0" for roll number/phone
            if val.endswith(".0") and val[:-2].replace("-", "").isdigit():
                val = val[:-2]
            item[canon_col] = val

        # If user provided separate First Name and Last Name
        if has_split_name:
            fname = item.get("first_name", "").strip()
            lname = item.get("last_name", "").strip()
            full = f"{fname} {lname}".strip()
            if full:
                item["name"] = full
                item["full_name"] = full

        # Ensure "full_name" and "name" are synchronized
        if "name" in item and "full_name" not in item:
            item["full_name"] = item["name"]
        elif "full_name" in item and "name" not in item:
            item["name"] = item["full_name"]

        # Clean email to lowercase
        if "email" in item and item["email"]:
            item["email"] = item["email"].lower().strip()

        normalized_rows.append(item)

    return normalized_rows


# ══════════════════════════════════════════════════════════════════════
# REFERENCE TEMPLATE GENERATORS
# ══════════════════════════════════════════════════════════════════════

SAMPLE_ROSTER_DATA = [
    {
        "Full Name": "John Doe",
        "Email Address": "john.doe@university.edu",
        "Roll Number": "CS-2026-001",
        "Phone Number": "+1 555-0101",
        "Division / Section": "Section A",
        "Department": "Computer Science & Engineering"
    },
    {
        "Full Name": "Jane Smith",
        "Email Address": "jane.smith@university.edu",
        "Roll Number": "CS-2026-002",
        "Phone Number": "+1 555-0102",
        "Division / Section": "Section A",
        "Department": "Computer Science & Engineering"
    },
    {
        "Full Name": "Alex Johnson",
        "Email Address": "alex.j@university.edu",
        "Roll Number": "CS-2026-003",
        "Phone Number": "+1 555-0103",
        "Division / Section": "Section B",
        "Department": "Data Science & AI"
    },
    {
        "Full Name": "Priya Patel",
        "Email Address": "priya.p@university.edu",
        "Roll Number": "CS-2026-004",
        "Phone Number": "+1 555-0104",
        "Division / Section": "Section B",
        "Department": "Artificial Intelligence"
    }
]


def generate_student_template_csv() -> str:
    """Generates standard UTF-8 CSV reference template."""
    output = io.StringIO()
    writer = csv.writer(output)
    headers = list(SAMPLE_ROSTER_DATA[0].keys())
    writer.writerow(headers)
    for row in SAMPLE_ROSTER_DATA:
        writer.writerow([row[h] for h in headers])
    return output.getvalue()


def generate_student_template_excel() -> bytes:
    """Generates professionally styled Excel (.xlsx) reference template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Roster Template"

    # Brand color styles
    header_fill = PatternFill(start_color="C84B18", end_color="C84B18", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="242321")
    thin_border = Border(
        left=Side(style="thin", color="E5E0D8"),
        right=Side(style="thin", color="E5E0D8"),
        top=Side(style="thin", color="E5E0D8"),
        bottom=Side(style="thin", color="E5E0D8")
    )

    headers = list(SAMPLE_ROSTER_DATA[0].keys())
    ws.append(headers)

    # Style Header Row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add Sample Rows
    for row_idx, row_dict in enumerate(SAMPLE_ROSTER_DATA, start=2):
        row_values = [row_dict[h] for h in headers]
        ws.append(row_values)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            # Center roll & phone & division
            if col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit Column Widths with padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
